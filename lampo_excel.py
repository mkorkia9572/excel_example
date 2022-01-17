import sys
from openpyxl import load_workbook

# avaa pohja.xlsx
wb = load_workbook("pohja.xlsx")

# ota käsittelyyn välilehti "Syote"
ws = wb["Syote"]

# käy läpi kuusi parametria (indeksit 1-6)
for i in range(1, 7):
    # luo excel solun viite
    cell = "B"+str(i+1)
    # syötä arvo soluun
    ws[cell] = float(sys.argv[i])
    # muokkaa solun muodoksi numero (on string by default)
    ws[cell].number_format = "0.0"

# tallenna excel
wb.save("lampotila.xlsx")